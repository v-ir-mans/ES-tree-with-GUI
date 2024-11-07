from csv import DictReader
from openpyxl import load_workbook
from Tree.Classes.EntryClasses import Entry_list, Entry

def read_csv_to_entries(path):
    entries = Entry_list()

    with open(path, mode='r', newline='', encoding='utf-8') as csvfile:
        reader = DictReader(csvfile)
        for row in reader:
            id_key, id=list(row.items())[0]
            klase_key, klase=list(row.items())[-1]

            other_columns_dict = {key: value for key, value in row.items() if key != id_key and key !=klase_key}
            full_row = row
            
            entry = Entry(klase, id, other_columns_dict, full_row)
            entries.append(entry)
    return entries


# ChatGPT 100%
def read_excel_to_entries(path, sheet_name=0):
    entries = Entry_list()
    
    # Load the workbook and select the sheet
    workbook = load_workbook(filename=path, data_only=True)
    if isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name]
    
    # Get header row as column names
    columns = [str(cell.value) for cell in sheet[1]]
    
    # Iterate over each row, starting from the second row (to skip headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Extract the first and last columns as id and klase, and convert them to strings
        id_key, id = columns[0], str(row[0])
        klase_key, klase = columns[-1], str(row[-1])
        
        # Create a dictionary of other columns except id and klase
        other_columns_dict = {columns[i]: str(row[i]) for i in range(len(row)) if columns[i] != id_key and columns[i] != klase_key}
        
        # Convert the entire row to a dictionary with all values as strings
        full_row = {columns[i]: str(row[i]) for i in range(len(row))}
        
        # Create and append an Entry object
        entry = Entry(klase, id, other_columns_dict, full_row)
        entries.append(entry)
    
    return entries